# report_generator.py

import pandas as pd
import logging
from openpyxl.styles import PatternFill
from datetime import datetime


class ReportGenerator:
    def __init__(self):
        self.bullish_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.bearish_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.bullish_terms = ['Bullish', 'Uptrend', 'Golden Cross', 'Oversold', 'Above SMA50']
        self.bearish_terms = ['Bearish', 'Downtrend', 'Death Cross', 'Overbought', 'Below SMA50']

    def _apply_conditional_formatting(self, worksheet, df):
        """Applies color formatting to signal and score cells."""
        header_row = {cell.value: cell.column for cell in worksheet[1]}

        for col_name, col_idx in header_row.items():
            if 'Signal' in str(col_name):
                for row_idx, cell_value in enumerate(df[col_name], 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if any(term in str(cell_value) for term in self.bullish_terms):
                        cell.fill = self.bullish_fill
                    elif any(term in str(cell_value) for term in self.bearish_terms):
                        cell.fill = self.bearish_fill

            if 'Score' in str(col_name):
                for row_idx, cell_value in enumerate(df[col_name], 2):
                    try:
                        numeric_value = float(cell_value)
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if numeric_value > 0:
                            cell.fill = self.bullish_fill
                        elif numeric_value < 0:
                            cell.fill = self.bearish_fill
                    except (ValueError, TypeError):
                        continue

    def generate_report(self, all_results, master_rankings=None, filename=None):
        """Generates a timestamped Excel report."""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            filename = f'stock_analysis_report_{timestamp}.xlsx'

        logging.info(f"Generating report: {filename}...")

        has_master_rankings = master_rankings is not None and len(master_rankings) > 0
        has_all_results = any(all_results.values())

        if not has_all_results and not has_master_rankings:
            logging.warning("No data was processed. The report will not be generated.")
            return

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            if has_master_rankings:
                summary_df = pd.DataFrame(master_rankings)

                # --- FIX: Define and apply the desired column order ---
                desired_order = ['Ticker', 'Short_Term_Score', 'Medium_Term_Score', 'Long_Term_Score', 'Master_Score']
                # Filter to only include columns that actually exist in the dataframe
                final_columns = [col for col in desired_order if col in summary_df.columns]
                summary_df = summary_df[final_columns]

                summary_df.set_index('Ticker', inplace=True)
                summary_df.to_excel(writer, sheet_name='Summary_Rankings')
                # --- FIX: Conditional formatting is now skipped for this sheet ---
                logging.info("Created 'Summary_Rankings' sheet.")

            if has_all_results:
                for timeframe_name, results in all_results.items():
                    if not results:
                        logging.warning(f"No results for {timeframe_name}. Skipping sheet.")
                        continue

                    df = pd.DataFrame(results).set_index('Ticker')
                    df.to_excel(writer, sheet_name=timeframe_name)
                    # Apply formatting to the detailed sheets
                    self._apply_conditional_formatting(writer.sheets[timeframe_name], df.reset_index())
                    logging.info(f"Wrote and formatted sheet for {timeframe_name}.")

        logging.info(f"Excel report saved as '{filename}'.")